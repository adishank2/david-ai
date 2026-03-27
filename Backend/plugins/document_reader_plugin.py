"""Document reader plugin for David AI Assistant.
Reads and summarizes PDF, Word, Excel, and text files.
"""

import os
from typing import Dict, List
from plugins.base import BasePlugin
from core.logger import get_logger
from core.config import FILE_OPS_ALLOWED_DIRS
from ai.llm import ask_llm

logger = get_logger(__name__)


class DocumentReaderPlugin(BasePlugin):
    """Read, extract text from, and summarize documents."""

    def __init__(self):
        super().__init__()
        self.userprofile = os.environ.get("USERPROFILE", "")
        self.allowed_paths = [
            os.path.join(self.userprofile, d.strip())
            for d in FILE_OPS_ALLOWED_DIRS
        ]

    def get_intents(self) -> List[str]:
        return ["read_document", "summarize_document"]

    def get_description(self) -> str:
        return "Read and summarize PDF, Word, Excel, and text documents"

    def get_prompt_examples(self) -> str:
        return """read_document:
{
  "intent": "read_document",
  "path": "Documents/report.pdf"
}

summarize_document:
{
  "intent": "summarize_document",
  "path": "Desktop/notes.docx"
}"""

    def execute(self, intent: Dict) -> str:
        intent_type = intent.get("intent")

        try:
            path = self._resolve_path(intent.get("path", ""))
            if not path:
                return "Please provide a file path."

            if not self._is_allowed(path):
                return "Access denied to this location."

            if not os.path.exists(path):
                return f"File not found: {os.path.basename(path)}"

            ext = os.path.splitext(path)[1].lower()

            if intent_type == "read_document":
                text = self._extract_text(path, ext)
                if not text:
                    return f"Could not read {os.path.basename(path)}. The file may be empty or in an unsupported format."

                # Return first 500 chars for speech
                preview = text[:500]
                if len(text) > 500:
                    preview += f"... ({len(text)} total characters)"

                return f"Contents of {os.path.basename(path)}: {preview}"

            elif intent_type == "summarize_document":
                text = self._extract_text(path, ext)
                if not text:
                    return f"Could not read {os.path.basename(path)}."

                # Truncate for LLM context window
                truncated = text[:3000]
                prompt = f"""Summarize the following document in 2-3 concise sentences:

{truncated}"""

                summary = ask_llm(prompt, temperature=0.3, num_predict=200)
                return f"Summary of {os.path.basename(path)}: {summary}"

            else:
                return "Unknown document command."

        except Exception as e:
            logger.error(f"Document reader error: {e}")
            return f"Error reading document: {e}"

    def _extract_text(self, path: str, ext: str) -> str:
        """Extract text content from various file formats."""

        # Plain text files
        if ext in [".txt", ".md", ".csv", ".log", ".json", ".xml",
                   ".py", ".js", ".html", ".css", ".yaml", ".yml",
                   ".ini", ".cfg", ".conf", ".bat", ".sh"]:
            return self._read_text(path)

        # PDF files
        elif ext == ".pdf":
            return self._read_pdf(path)

        # Word documents
        elif ext in [".docx", ".doc"]:
            return self._read_docx(path)

        # Excel files
        elif ext in [".xlsx", ".xls"]:
            return self._read_excel(path)

        else:
            return ""

    @staticmethod
    def _read_text(path: str) -> str:
        """Read plain text file."""
        try:
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                return f.read()
        except Exception as e:
            logger.error(f"Text read error: {e}")
            return ""

    @staticmethod
    def _read_pdf(path: str) -> str:
        """Read PDF file using PyPDF2."""
        try:
            from PyPDF2 import PdfReader

            reader = PdfReader(path)
            text_parts = []
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text_parts.append(page_text)

            text = "\n".join(text_parts)
            logger.info(f"PDF read: {len(reader.pages)} pages, {len(text)} chars")
            return text

        except ImportError:
            logger.error("PyPDF2 not installed. Run: pip install PyPDF2")
            return "[PDF reading requires PyPDF2. Install with: pip install PyPDF2]"
        except Exception as e:
            logger.error(f"PDF read error: {e}")
            return ""

    @staticmethod
    def _read_docx(path: str) -> str:
        """Read Word document using python-docx."""
        try:
            from docx import Document

            doc = Document(path)
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            text = "\n".join(paragraphs)
            logger.info(f"DOCX read: {len(paragraphs)} paragraphs, {len(text)} chars")
            return text

        except ImportError:
            logger.error("python-docx not installed. Run: pip install python-docx")
            return "[Word reading requires python-docx. Install with: pip install python-docx]"
        except Exception as e:
            logger.error(f"DOCX read error: {e}")
            return ""

    @staticmethod
    def _read_excel(path: str) -> str:
        """Read Excel file using openpyxl."""
        try:
            from openpyxl import load_workbook

            wb = load_workbook(path, read_only=True, data_only=True)
            text_parts = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                text_parts.append(f"--- Sheet: {sheet_name} ---")
                row_count = 0
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    if any(cells):
                        text_parts.append(" | ".join(cells))
                        row_count += 1
                    if row_count >= 50:  # Limit rows
                        text_parts.append(f"... (showing first 50 rows)")
                        break

            wb.close()
            text = "\n".join(text_parts)
            logger.info(f"Excel read: {len(wb.sheetnames)} sheets, {len(text)} chars")
            return text

        except ImportError:
            logger.error("openpyxl not installed. Run: pip install openpyxl")
            return "[Excel reading requires openpyxl. Install with: pip install openpyxl]"
        except Exception as e:
            logger.error(f"Excel read error: {e}")
            return ""

    def _resolve_path(self, path: str) -> str:
        """Resolve relative path to absolute."""
        if not path:
            return ""
        if os.path.isabs(path):
            return path
        for allowed in self.allowed_paths:
            if os.path.basename(allowed).lower() in path.lower():
                relative = path.split("/", 1)[-1] if "/" in path else path.split("\\", 1)[-1] if "\\" in path else ""
                return os.path.join(allowed, relative) if relative else allowed
        return os.path.join(self.userprofile, "Desktop", path)

    def _is_allowed(self, path: str) -> bool:
        """Check if path is in allowed directories."""
        if not path:
            return False
        abs_path = os.path.abspath(path)
        for allowed in self.allowed_paths:
            if abs_path.startswith(os.path.abspath(allowed)):
                return True
        return False
